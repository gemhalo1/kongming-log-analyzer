from openpyxl import Workbook
from openpyxl.styles import Font
from openpyxl.utils import get_column_letter

from typing import List
from .model import DialogRound
from .constants import CLEAN_CONTEXT_MAGIC_STRING
from .utils import convert_timestamp, calculate_time_difference

def set_column_width(ws, column_index, width):
    c = get_column_letter(column_index)
    ws.column_dimensions[c].width = width

def print_dialog_round_to_excel(rounds: List[DialogRound], filename: str):

    wb = Workbook()
    ws = wb.active
    ws.title = "对话记录"

    titles = ["时间戳", 
              "trace_id", 
              "眼镜类型",
              "位置",
              "originType",
              "functionType",
              "locale",
              "时区",
              "语种",
              "首轮",

              "NLU查询",
              "NLU意图",
              "NLU回答",
              "NLU Error",
              "NLU耗时",
              "LLM查询",
              "LLM意图",
              "LLM场景",  # channel_type flag
              "图像文件",
              "角色扮演",
              "深度思考",
              "深度搜索",
              "视觉辅助",
              "清上下文", # clean_context flag
              "LLM回答",
              "LLM思考",
              "LLM搜索数据",
              "LLM状态",
              "LLM耗时",

              "deviceId", 
              "glassDeviceId",
              'iotDeviceId',
              "accountId",
              "xjAccountId",
              'sessionId',
              'msgId',

            ]
    ws.append(titles)

    title_font_attribs = {
        'name' : 'Calibri',
        'charset': None,
        'family' : 2,
        'b' : True,
        'i' : False,
        'strike' : False,
        'outline' : False,
        'shadow' : False,
        # 'condense' : '',
        'color' : None,
        # 'extend' : '',
        'sz' : 10,
        'u' : None,   # 'double', 'doubleAccounting', 'single', 'singleAccounting'
        # 'vertAlign' : '',
        'scheme':"minor"
    }

    title_font = Font(**title_font_attribs)

    content_font_attribs = {
        'name' : 'Courier New',
        'charset': None,
        'family' : 3,
        'b' : False,
        'i' : False,
        'strike' : False,
        'outline' : False,
        'shadow' : False,
        # 'condense' : '',
        'color' : None,
        # 'extend' : '',
        'sz' : 8,
        'u' : None,   # 'double', 'doubleAccounting', 'single', 'singleAccounting'
        # 'vertAlign' : '',
        'scheme':"minor"
    }

    content_font = Font(**content_font_attribs)

    for col in range(len(titles)):
        ws.cell(1, col + 1).font = title_font

    set_column_width(ws, 1, 19)
    set_column_width(ws, 2, 28)
    set_column_width(ws, 4, 16)
    set_column_width(ws, 10, 12)
    set_column_width(ws, 11, 30)
    set_column_width(ws, 12, 25)
    set_column_width(ws, 13, 30)
    set_column_width(ws, 16, 30)
    set_column_width(ws, 19, 30)
    set_column_width(ws, 25, 30)
    set_column_width(ws, 33, 10)
    set_column_width(ws, 34, 25)
    set_column_width(ws, 36, 32)

    for idx in [30, 31, 32, 35]:
        set_column_width(ws, idx, 28)

    # 添加行数据
    for idx, round in enumerate(rounds, 1):  # 序号从1开始
        ws.append([
            # convert_timestamp(round.nlp_round.request_timestamp) if round.nlp_round.request_timestamp else "",
            round.nlp_round.request_timestamp,
            round.traceId or "",
            int(round.glassProduct) if round.glassProduct else "",
            str(round.location) if round.location else "",

            round.originType if round.originType is not None else "",
            round.functionType if round.functionType is not None else "",
            round.local or "",
            round.timeZone or "",
            round.nluLanguage or "",
            round.sessionFirstFlag if round.sessionFirstFlag is not None else "" or "",

            (round.nlp_round.query if round.nlp_round.query != CLEAN_CONTEXT_MAGIC_STRING else "<清除上下文>") if round.nlp_round else "",
            str(round.nlp_round.intent or "") if round.nlp_round else "",
            str(round.nlp_round.utterance) if round.nlp_round and round.nlp_round.utterance else "",
            str(round.nlp_round.error) if round.nlp_round and round.nlp_round.error else "",
            calculate_time_difference(round.nlp_round.request_timestamp, round.nlp_round.response_timestamp) if round.nlp_round and round.nlp_round.response_timestamp else "",

            (round.llm_round.query or "") if round.llm_round and round.llm_round.query else "",
            (round.llm_round.intent_name or "") if round.llm_round and round.llm_round.intent_name else "",
            round.llm_round.channel_type if round.llm_round and round.llm_round.channel_type is not None else "",

            '\n'.join([file.ossUrl for file in round.llm_round.files]) if round.llm_round and round.llm_round.files else "",

            round.llm_round.play_status if round.llm_round and round.llm_round.play_status is not None else "",
            round.llm_round.use_deepseek if round.llm_round and round.llm_round.use_deepseek is not None else "",
            round.llm_round.use_search if round.llm_round and round.llm_round.use_search is not None else "",
            round.llm_round.visual_aids_status if round.llm_round and round.llm_round.visual_aids_status is not None else "",
            round.llm_round.clean_context if round.llm_round and round.llm_round.clean_context is not None else "",

            (round.llm_round.answer or "") if round.llm_round and round.llm_round.answer else "",
            (round.llm_round.reason or "") if round.llm_round and round.llm_round.reason else "",
            str(round.llm_round.thoughts_data) if round.llm_round and round.llm_round.thoughts_data else "",
            round.llm_round.base_status if round.llm_round is not None else "",
            calculate_time_difference(round.llm_round.request_timestamp, round.llm_round.response_timestamp) if round.llm_round and round.llm_round.response_timestamp else "",

            round.deviceId or "",
            round.glassDeviceId or "",
            round.iotDeviceId or "",
            round.accountId or "",
            round.xjAccountId or "",
            round.sessionId or "",
            round.msgId or "",
        ])

        for col in range(len(titles)):
            ws.cell(idx + 1, col + 1).font = content_font
    
    wb.save(filename)
